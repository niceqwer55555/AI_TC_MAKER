from openpyxl import load_workbook  # 读取excel中的数据
# from Autotestplat import settings  # 获取需要读取“测试数据”的路径
from copy import deepcopy
import re

class OperateExcel():
    '''
    1、read_data()，遍历每一行数据并放到List,并返回list，使用列表套列表的形式，每一行为一个内层列表，无数行存储在一个大列表中
    2、output_data_dict()，将数据中的内层列表转换为字典存储，外层依然是列表
    3、test_data()，将AssertInfo,SendData等字段对应的值转换为 dict 形式
    '''
    def __init__(self,file_name):
        '''
        初始化，为读取excel做准备
        '''
        # self.filePath= globalConfig.excel_data_path + file_name
        self.filePath= file_name


        self.wb = load_workbook(self.filePath)
        # 获取所有sheet页名字
        self.SheetNames = self.wb.sheetnames
        self.sheet = self.wb[self.SheetNames[0]]  # self.wb[SheetName]
        self.MaxRowNum = self.sheet.max_row
        self.MaxColNum=self.sheet.max_column


    def read_rows_data(self,row):
        '''
        2、遍历每一行数据并放到List,并返回list，使用列表套列表的形式，每一行为一个内层列表，无数行存储在一个大列表中
        :return:
        '''
        dataList = []
        for line in list(self.sheet.rows)[row-1:row]:  # 获取excel中每一行的数据。rows这个函数 用来按行读取；columns这个函数 用来按列读取。
            tmpList = []
            for i in range(0, self.MaxColNum):  # 指定获取每一行中的1-12列的数据，不需要 12列后面的数据。
                tmpList.append(line[i].value)
            dataList.append(tmpList)
        return dataList

    def read_columns_data(self):
        '''
        2、遍历每一行数据并放到List,并返回list，使用列表套列表的形式，每一行为一个内层列表，无数行存储在一个大列表中
        :return:
        '''
        dataList = []
        for line in list(self.sheet.columns)[0:]:  # 获取excel中每一行的数据。rows这个函数 用来按行读取；columns这个函数 用来按列读取。
            tmpList = []
            for i in range(1, self.MaxRowNum):  # 指定获取每一行中的1-12列的数据，不需要 12列后面的数据。
                tmpList.append(line[i].value)
            dataList.append(tmpList)
        return dataList

    def write_to_excel(self,path: str, col_index,row_index, row_item):
        # 激活一个sheet
        sheet = self.wb.active
        # 为sheet设置一个title
        # sheet.title = sheetStr
        # 添加表头（不需要表头可以不用加）
        # data.insert(0, list(info))
        # 开始遍历数组
        # for row_index, row_item in enumerate(data):
        #     for col_index, col_item in enumerate(row_item):
        #         # 写入
        #         sheet.cell(row=row_index + 1, column=col_index + 1, value=col_item)
        row_item_str = '\n'.join(row_item)
        sheet.cell(row=row_index, column=col_index, value=row_item_str)
        # 写入excel文件 如果path路径的文件不存在那么就会自动创建
        self.wb.save(path)
        # print('写入成功')

    def find_column_value(self,target_value):
        sheet = self.wb.active
        column_list = []
        row_list = []
        for column in sheet.iter_cols(min_row=2,max_row=None,min_col=2,max_col=None,):
            for cell in column:
                if cell.value == target_value:
                    # print(cell.row)
                    column_list.append(cell.column)
                    row_list.append(cell.row)
        return row_list
        # return None


    def output_data_dict(self,row):
        '''
        3、将数据中的内层列表转换为字典存储，外层依然是列表。并返回
        形如：
        [
            {'me':'zk','you':'Bill'},
            {'me1':'zk1','you2':'Bill2'}
        ]
        :return:外列表，内字典。内字典中的 ’send_data‘对应的value值是字符串，需要使用 eval()函数 将其转换成字典才能使用。
        '''
        #print('最大行数：', self.MaxRowNum,'\n最大列数：',self.MaxColNum)
        readDataList = []
        # read_rows_data_list = self.read_rows_data(2)
        # read_columns_data_list = self.read_columns_data()
        # pprint(read_data_list)
        # for i in range(1, len(read_rows_data_list)):  # 外层列表循环
        temDict = {}
        for j in range(0, len(self.read_rows_data(1)[0])):  # 内层列表循环
            '''
            下面的字典赋值语句，将excel中的第一行中的字段名，与每一列中的数据一一对应起来，形成key-value这样的字典。
            '''
            # print(read_data_list[0])
            # print(read_data_list[i])
            # dictionary = dict(zip(read_rows__data_list[0], read_rows__data_list[i]))
            # print(dictionary)
            temDict[self.read_rows_data(1)[0][j]] = self.read_rows_data(row)[0][j]
        readDataList.append(temDict)  # 将每一行数据 与 字段名 形成的 字典 作为 列表的值 存储起来。
        # print(readDataList)
        return readDataList

    def test_data(self,row):
        '''
        4、将AssertInfo,SendData等字段对应的值转换为 dict 形式

        *** 关于正则表达式的解释 ***

            re.search("\{.*\}",list['SendData'],re.S)!=None
            解释：
            （1）\{.*\} ，表示在母串中匹配 “{} 中有任意字符” 的子串；\{，表示转义，因为{}在正则表达式中有特殊含义；.* 表示任意字符（通用匹配）
            （2）list['SendData'] 母串
            （3）re.S ，表示使“.”匹配包括换行在内的所有字符；“.” 表示匹配任意字符（除换行符\n）
            （4）当search()有结果时，会返回相应的对象结果；如果没有结果的话，那么就会返回 None

        :return: test_data，作为读取的最终输出；数据结构 [{{}},{{}}]
        '''
        readDataList=self.output_data_dict(row)
        test_data=deepcopy(readDataList)

        for i in range(0,len(readDataList)):
            list=readDataList[i]

            '''
                下面这些if判断都是实现的同样功能
                （1）当某值不为 None 
                （2）当该值中存在 “{任意字符}” 这样的字符；详情请看方法注释
                （3）同时符合（1）和（2）中条件时则执行if下面的代码
                （4）将该值使用 eval() 函数进行转换后，并对相应的key重新赋值
            '''
            for key in list.keys():
                if re.search("\{.*\}",str(list[key]),re.S)!=None:
                    field_data=eval(list[key])
                    test_data[i][key]=field_data
                elif re.search("\[.*\]",str(list[key]),re.S)!=None:
                    field_data = eval(list[key])
                    test_data[i][key] = field_data

        return test_data



if __name__=="__main__":
    import json
    test=OperateExcel('test.xlsx')
    data_all=test.test_data(3)
    # print(data_all)
    # read_columns_data_list = test.read_columns_data()
    # print(read_columns_data_list)
    # print(data_all[0])
    # print(list(data_all[0].items()))
    # print(type(data_all))
    # print(type(data_all[0]))
    # result = []
    # for i in range(0, len(data_all)):
    #     # print(data_all[i])
    #     result += list(data_all[i].items())
    # print(result)
    # result = tuple()
    # lst = [(1,), (2, 3), (5, 6)]
    # print(type(lst))
    # for i in data_all:
    #     result = result.__add__(i)
    # result = tuple()
    # for i in data_all[0]:
    #     result = result.__add__(i)
    # print(json.dumps(data_all,indent=4,ensure_ascii=False)) # 按照json格式输出
    # print(type(data_all))
    # print(type(data_all[0]))

    # 数据结构1 path 文件的路径
    # path = test.filePath
    # # 数据结构1Excel 中sheet 的名字
    # sheetStr = 'testcase'
    #
    # info = ['None']
    # # 数据结构1数据
    # writeData = [['1711198647']]
    #
    # # 执行
    # test.write_to_excel(path, sheetStr, info, writeData)

    # test.find_column_value('1')
    # 调用函数查找特定值所在的列值
    # column = test.find_column_value(1)
    # print(f"特定值所在的列值为: {column}")
    # # dd = test.read_rows_data(4)
    # print(dd)
    cc = ['1. 打开包含列表的页面', '2. 检查列表展示的顺序', '3. 查看初始分页控件的显示', '4. 修改分页控件的每页显示条目为20条', '5. 点击分页控件中的下一页', '6. 输入一个页码以跳转到特定页面', '7. 点击分页控件中的首页按钮', '8. 点击分页控件中的尾页按钮']
    cc_str = '\n'.join(cc)
    test.write_to_excel('test.xlsx',8,2,cc_str)

